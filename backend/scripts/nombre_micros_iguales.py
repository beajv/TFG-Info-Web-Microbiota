import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from collections import defaultdict
import colorsys

# === PARTE 1: leer y ordenar por nombre_limpio manteniendo texto original ===
archivo_entrada = "microorganismos_detectados_limpio.xlsx"
archivo_salida = "microorganismos_coloreado.xlsx"

# Cargar el Excel tal cual (sin cambiar nombres)
df = pd.read_excel(archivo_entrada)

# Ordenar por nombre_limpio tal como está (sin minúsculas ni modificaciones)
df_sorted = df.sort_values(by="nombre_limpio", kind="stable")

# Guardar en Excel (sin colores aún)
df_sorted.to_excel(archivo_salida, index=False)

# === PARTE 2: aplicar colores a nombres repetidos ===
# Detectar duplicados exactos
repetidos = df_sorted["nombre_limpio"].value_counts()
nombres_repetidos = repetidos[repetidos > 1].index.tolist()

# Generar colores únicos (HSV → RGB)
def generar_color(i, total):
    hue = i / total
    r, g, b = colorsys.hsv_to_rgb(hue, 0.4, 1)
    return PatternFill(start_color=f'{int(r*255):02X}{int(g*255):02X}{int(b*255):02X}',
                       end_color=f'{int(r*255):02X}{int(g*255):02X}{int(b*255):02X}',
                       fill_type="solid")

# Asignar colores
colores = {name: generar_color(i, len(nombres_repetidos)) for i, name in enumerate(nombres_repetidos)}

# Cargar Excel con openpyxl
wb = load_workbook(archivo_salida)
ws = wb.active

# Detectar columna "nombre_limpio"
col_nombre = None
for idx, cell in enumerate(ws[1], 1):
    if cell.value == "nombre_limpio":
        col_nombre = idx
        break

# Pintar celdas con color si están repetidas
if col_nombre:
    for row in range(2, ws.max_row + 1):
        valor = ws.cell(row=row, column=col_nombre).value
        if valor in colores:
            ws.cell(row=row, column=col_nombre).fill = colores[valor]

wb.save(archivo_salida)
print(" Listo: microorganismos_coloreado.xlsx generado con colores por duplicado.")
